"""
ShipTrack Telegram Bot
Топик "Приёмка"  → отгрузки (АЛИ / МКО)
Топик "Отгрузка" → возвраты (xlsx, xls, csv, pdf)
Другие топики    → молчит
Личка            → работает по подписи
"""
import os, re, csv, logging, tempfile, threading, time, asyncio
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
import openpyxl
import pdfplumber
from supabase import create_client
from telegram import Update
from telegram.ext import (
    Application, MessageHandler, CommandHandler,
    filters, ContextTypes
)

# ══════════════════════════════════════════════════════════════
# ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ — задаются в Render → Environment
BOT_TOKEN    = os.environ.get("BOT_TOKEN", "")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

# ID топиков — узнайте через /topicid в каждом топике
TOPIC_PRIEMKA  = int(os.environ.get("TOPIC_PRIEMKA", "0"))
TOPIC_OTGRUZKA = int(os.environ.get("TOPIC_OTGRUZKA", "0"))
# ══════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s"
)
log = logging.getLogger(__name__)


# ── Health check сервер для Render ────────────────────────────────────────────

class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"OK")

    def log_message(self, format, *args):
        pass


_health_port   = int(os.environ.get("PORT", 8000))
_health_server = HTTPServer(("0.0.0.0", _health_port), HealthHandler)
threading.Thread(target=_health_server.serve_forever, daemon=True).start()
log.info("✅ Health server запущен на порту %s", _health_port)


# ── Самопинг — не даёт Render усыпить сервис ─────────────────────────────────

def self_ping():
    import urllib.request
    public_url = os.environ.get("PUBLIC_URL", "")
    if not public_url:
        log.warning("⚠️ PUBLIC_URL не задан — самопинг отключён")
        return
    while True:
        time.sleep(4 * 60)
        try:
            urllib.request.urlopen(public_url, timeout=10)
            log.info("🏓 Self-ping OK")
        except Exception as e:
            log.warning("🏓 Self-ping failed: %s", e)

threading.Thread(target=self_ping, daemon=True).start()


# ── Инициализация Supabase ────────────────────────────────────────────────────

db = create_client(SUPABASE_URL, SUPABASE_KEY)

# Память о сигналах возврата: {chat_id: datetime}
return_signals = {}
RETURN_WINDOW  = 5 * 60  # 5 минут
RETURN_WORDS   = {"возврат", "возвратов", "возврата", "возвраты"}


# ── Вспомогательные функции ───────────────────────────────────────────────────

def has_return_word(text):
    if not text:
        return False
    words = re.findall(r"[а-яёa-z]+", text.lower())
    return any(w in RETURN_WORDS for w in words)


def recent_return_signal(chat_id):
    ts = return_signals.get(chat_id)
    if not ts:
        return False
    return (datetime.now() - ts).total_seconds() <= RETURN_WINDOW


def parse_date(text):
    today = datetime.now()
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?", text)
    if m:
        d, mo = m.group(1).zfill(2), m.group(2).zfill(2)
        y = m.group(3) or str(today.year)
        if len(y) == 2:
            y = "20" + y
        try:
            dt = datetime(int(y), int(mo), int(d))
            return dt.strftime("%Y-%m-%d"), f"{d}.{mo}"
        except ValueError:
            pass
    return today.strftime("%Y-%m-%d"), today.strftime("%d.%m")


def detect(caption, chat_id):
    today = datetime.now()
    up = (caption or "").strip().upper()

    if any(k in up for k in ("ВОЗВРАТ", "VOZVRAT", "RETURN")):
        di, dl = parse_date(up)
        return "return", "ali", di, dl

    if recent_return_signal(chat_id):
        log.info("🔄 Файл = возврат (по недавнему сообщению)")
        return "return", "ali", today.strftime("%Y-%m-%d"), today.strftime("%d.%m")

    if "АЛИ" in up or "ALI" in up:
        di, dl = parse_date(up)
        return "shipment", "ali", di, dl

    if "МКО" in up or "MKO" in up:
        di, dl = parse_date(up)
        return "shipment", "mko", di, dl

    return "unknown", None, today.strftime("%Y-%m-%d"), today.strftime("%d.%m")


def act_already_exists(act_number: str) -> bool:
    """Проверяет есть ли уже возврат с таким номером акта в базе."""
    if not act_number:
        return False
    result = db.table("returns").select("id").eq("act_number", act_number).execute()
    return len(result.data) > 0


# ── Парсеры отгрузок ──────────────────────────────────────────────────────────

def parse_shipment_csv(path):
    """CSV файл отгрузки — колонка Штрихкод."""
    parcels = []
    for enc in ["utf-8-sig", "utf-8", "cp1251"]:
        try:
            with open(path, "r", encoding=enc) as f:
                s = f.read(2048)
                f.seek(0)
                delim = ";" if s.count(";") > s.count(",") else ","
                for row in csv.DictReader(f, delimiter=delim):
                    barcode = name = phone = address = ""
                    weight = 0
                    for k, v in row.items():
                        if not k:
                            continue
                        kl = k.lower()
                        if "штрихкод" in kl or "barcode" in kl:
                            barcode = str(v or "").strip()
                        elif "имя" in kl or "name" in kl or "фио" in kl:
                            name = str(v or "").strip()
                        elif "телефон" in kl or "phone" in kl:
                            phone = str(v or "").strip()
                        elif "адрес" in kl or "address" in kl:
                            address = str(v or "").strip()
                        elif "вес" in kl or "weight" in kl:
                            try:
                                weight = int(float(str(v or 0)))
                            except Exception:
                                pass
                    if not barcode and row:
                        barcode = str(list(row.values())[-1] or "").strip()
                    if barcode and len(barcode) >= 5:
                        parcels.append({
                            "barcode":        barcode,
                            "recipient_name": name,
                            "phone":          phone,
                            "address":        address,
                            "weight":         weight,
                        })
            if parcels:
                return parcels
        except Exception:
            continue
    return parcels


def parse_shipment_excel(path):
    """Excel файл отгрузки — колонка Штрихкод."""
    wb = openpyxl.load_workbook(path, data_only=True)
    parcels = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c or "").lower() for c in rows[0]]
        bi = next((i for i, h in enumerate(header) if "штрихкод" in h or "barcode" in h), len(header) - 1)
        ni = next((i for i, h in enumerate(header) if "имя" in h or "name" in h or "фио" in h), -1)
        pi = next((i for i, h in enumerate(header) if "телефон" in h or "phone" in h), -1)
        ai = next((i for i, h in enumerate(header) if "адрес" in h or "address" in h), -1)
        wi = next((i for i, h in enumerate(header) if "вес" in h or "weight" in h), -1)

        def g(row, i):
            return str(row[i] or "").strip() if 0 <= i < len(row) else ""

        for row in rows[1:]:
            barcode = g(row, bi)
            if barcode and len(barcode) >= 5:
                w = 0
                if wi >= 0:
                    try:
                        w = int(float(str(row[wi] or 0)))
                    except Exception:
                        pass
                parcels.append({
                    "barcode":        barcode,
                    "recipient_name": g(row, ni),
                    "phone":          g(row, pi),
                    "address":        g(row, ai),
                    "weight":         w,
                })
    return parcels


# ── Парсеры возвратов ─────────────────────────────────────────────────────────

def parse_return_csv(path):
    """CSV файл возврата — колонки: Номер заказа, Кол-во, Стоимость."""
    items = []
    act_number = ""
    for enc in ["utf-8-sig", "utf-8", "cp1251"]:
        try:
            with open(path, "r", encoding=enc) as f:
                s = f.read(2048)
                f.seek(0)
                delim = ";" if s.count(";") > s.count(",") else ","
                for row in csv.DictReader(f, delimiter=delim):
                    order = ""
                    qty   = 1
                    cost  = 0.0
                    for k, v in row.items():
                        if not k:
                            continue
                        kl = k.lower()
                        if "номер" in kl or "order" in kl:
                            order = str(v or "").strip()
                        elif "кол" in kl or "шт" in kl or "qty" in kl:
                            try:
                                qty = int(v or 1)
                            except Exception:
                                pass
                        elif "стоимость" in kl or "сумм" in kl or "cost" in kl:
                            try:
                                cost = float(str(v or 0).replace(" ", ""))
                            except Exception:
                                pass
                    if order and len(order) >= 5:
                        items.append({"order_number": order, "quantity": qty, "cost": cost})
            if items:
                return items, act_number
        except Exception:
            continue
    return items, act_number


def parse_return_excel(path):
    """Excel файл возврата — Акт приема-передачи."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    items = []
    act_number = ""

    for row in rows[:5]:
        for cell in row:
            if cell and "акт" in str(cell).lower():
                m = re.search(r"№\s*(\d+)", str(cell))
                if m:
                    act_number = m.group(1)

    header_idx = None
    for i, row in enumerate(rows):
        s = " ".join(str(c or "").lower() for c in row)
        if "номер заказа" in s or "номер" in s:
            header_idx = i
            break

    if header_idx is None:
        return items, act_number

    header = [str(c or "").lower() for c in rows[header_idx]]
    oi = next((i for i, h in enumerate(header) if "номер" in h), 1)
    qi = next((i for i, h in enumerate(header) if "кол" in h or "шт" in h), 2)
    ci = next((i for i, h in enumerate(header) if "стоимость" in h or "сумм" in h), 3)

    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        try:
            int(row[0])
        except (TypeError, ValueError):
            continue
        order = str(row[oi] or "").strip() if oi < len(row) else ""
        if not order or len(order) < 5:
            continue
        qty = 1
        cost = 0.0
        try:
            qty = int(row[qi] or 1)
        except Exception:
            pass
        try:
            cost = float(row[ci] or 0)
        except Exception:
            pass
        items.append({"order_number": order, "quantity": qty, "cost": cost})

    return items, act_number


def parse_return_pdf(path):
    """PDF файл возврата — Акт приема-передачи Uzum Market."""
    items = []
    act_number = ""

    with pdfplumber.open(path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

        m = re.search(r"Акт приема-передачи\s*№\s*(\d+)", full_text)
        if m:
            act_number = m.group(1)

        ORDER_RE = re.compile(
            r"^(\d+)\s+([A-Z]{2}\d+UZ)\s+(\d+)\s+([\d\s]+)$"
        )

        for line in full_text.splitlines():
            line = line.strip()
            if not line:
                continue
            m = ORDER_RE.match(line)
            if m:
                order    = m.group(2)
                qty      = int(m.group(3))
                cost_str = m.group(4).replace(" ", "")
                try:
                    cost = float(cost_str)
                except Exception:
                    cost = 0.0
                items.append({"order_number": order, "quantity": qty, "cost": cost})
                continue

            m2 = re.search(r"([A-Z]{2}\d+UZ)\s+(\d+)\s+([\d]+)", line)
            if m2 and not any(i["order_number"] == m2.group(1) for i in items):
                order = m2.group(1)
                qty   = int(m2.group(2))
                try:
                    cost = float(m2.group(3))
                except Exception:
                    cost = 0.0
                items.append({"order_number": order, "quantity": qty, "cost": cost})

    log.info("📄 PDF: акт №%s, %d заказов", act_number, len(items))
    return items, act_number


# ── Сохранение в Supabase ─────────────────────────────────────────────────────

def save_shipment(parcels, filename, sender, project, date, date_label):
    prefix = {"ali": "ALI", "mko": "UCB"}.get(project, "SHP")
    names  = {
        "ali": f"AliExpress от {date_label}",
        "mko": f"Uzum Crossborder от {date_label}",
    }
    name = names.get(project, f"Партия от {date_label}")

    base_id  = f"{prefix} {date_label}"
    existing = db.table("shipments").select("id").eq("project", project).eq("date", date).execute().data
    sid      = base_id if len(existing) == 0 else f"{base_id}-{len(existing) + 1}"

    db.table("shipments").insert({
        "id":            sid,
        "name":          name,
        "project":       project,
        "date":          date,
        "parcels_count": len(parcels),
        "status":        "new",
        "confirmed_at":  None,
        "note":          "",
        "filename":      filename,
        "sender":        sender,
    }).execute()
    for i in range(0, len(parcels), 50):
        db.table("parcels").insert(
            [{"shipment_id": sid, **p} for p in parcels[i:i + 50]]
        ).execute()
    return sid, name


def save_return(items, act_number, filename, sender, date, date_label):
    act_label = f" (Акт №{act_number})" if act_number else ""
    name      = f"Возврат от {date_label}{act_label}"
    total     = sum(i["cost"] for i in items)

    base_id  = f"RET {date_label}"
    existing = db.table("returns").select("id").eq("date", date).execute().data
    rid      = base_id if len(existing) == 0 else f"{base_id}-{len(existing) + 1}"

    db.table("returns").insert({
        "id":           rid,
        "name":         name,
        "date":         date,
        "orders_count": len(items),
        "total_cost":   total,
        "status":       "new",
        "confirmed_at": None,
        "note":         "",
        "filename":     filename,
        "act_number":   act_number,
        "sender":       sender,
    }).execute()
    for i in range(0, len(items), 50):
        db.table("return_items").insert(
            [{"return_id": rid, **item} for item in items[i:i + 50]]
        ).execute()
    return rid, name, total


# ── Обработчики Telegram ──────────────────────────────────────────────────────

async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text    = update.message.text or ""
    chat_id = update.message.chat_id
    if has_return_word(text):
        return_signals[chat_id] = datetime.now()
        log.info("🔔 Сигнал возврата в чате %s", chat_id)


async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc       = update.message.document
    fname     = doc.file_name or ""
    caption   = update.message.caption or ""
    chat_id   = update.message.chat_id
    thread_id = update.message.message_thread_id or 0
    is_group  = update.message.chat.type in ("group", "supergroup")

    ext = os.path.splitext(fname)[1].lower()
    if ext not in (".xlsx", ".xls", ".csv", ".pdf"):
        return

    sender = update.message.from_user.full_name or "Unknown"

    if TOPIC_PRIEMKA and TOPIC_OTGRUZKA:
        if thread_id == TOPIC_PRIEMKA:
            ftype, project, date, date_label = detect(caption, chat_id)
            if ftype == "unknown":
                await update.message.reply_text(
                    "❓ Добавьте подпись к файлу:\n"
                    "• `АЛИ 15.04` — AliExpress\n"
                    "• `МКО` — Uzum Crossborder",
                    parse_mode="Markdown"
                )
                return
        elif thread_id == TOPIC_OTGRUZKA:
            _, _, date, date_label = detect(caption, chat_id)
            ftype, project = "return", "ali"
        else:
            if is_group:
                return
            ftype, project, date, date_label = detect(caption, chat_id)
    else:
        ftype, project, date, date_label = detect(caption, chat_id)

    labels = {"shipment": "Отгрузка", "return": "Возврат", "unknown": "Файл"}
    msg = await update.message.reply_text(
        f"📊 Читаю *{fname}*...\nТип: *{labels[ftype]}*",
        parse_mode="Markdown"
    )

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        path = tmp.name

    try:
        await (await doc.get_file()).download_to_drive(path)

        # ── Возврат ──
        if ftype == "return":
            if ext == ".pdf":
                items, act_num = parse_return_pdf(path)
            elif ext == ".csv":
                items, act_num = parse_return_csv(path)
            else:
                items, act_num = parse_return_excel(path)

            if not items:
                await msg.edit_text("⚠️ Не нашёл данных о заказах в файле.")
                return

            # Проверка дублей по номеру акта
            if act_num and act_already_exists(act_num):
                log.info("⚠️ Акт №%s уже существует — пропускаем", act_num)
                await msg.edit_text(
                    f"⚠️ *Акт №{act_num} уже добавлен в систему.*\n\n"
                    f"Файл пропущен чтобы избежать дублей.",
                    parse_mode="Markdown"
                )
                return

            rid, name, total = save_return(items, act_num, fname, sender, date, date_label)
            await msg.edit_text(
                f"🔄 *Возврат добавлен!*\n\n"
                f"📋 {name}\n"
                f"📦 Заказов: *{len(items)}*\n"
                f"💰 Сумма: *{total:,.0f} ₽*\n"
                f"🆔 ID: `{rid}`",
                parse_mode="Markdown"
            )

        # ── Отгрузка ──
        elif ftype == "shipment":
            if ext == ".csv":
                parcels = parse_shipment_csv(path)
            else:
                parcels = parse_shipment_excel(path)

            if not parcels:
                await msg.edit_text(
                    "⚠️ Не нашёл штрихкодов. Нужна колонка *Штрихкод*.",
                    parse_mode="Markdown"
                )
                return
            proj_name = {
                "ali": "AliExpress 🛒",
                "mko": "Uzum Crossborder 📦",
            }.get(project, "")
            sid, name = save_shipment(parcels, fname, sender, project, date, date_label)
            await msg.edit_text(
                f"✅ *Отгрузка добавлена!*\n\n"
                f"{proj_name}\n"
                f"📋 {name}\n"
                f"📦 Посылок: *{len(parcels)}*\n"
                f"🆔 ID: `{sid}`",
                parse_mode="Markdown"
            )

        # ── Неизвестный тип ──
        else:
            await msg.edit_text(
                "❓ Не могу определить тип файла.\n\n"
                "Добавьте подпись к файлу:\n"
                "• `АЛИ 15.04` — отгрузка AliExpress\n"
                "• `МКО` — отгрузка Uzum\n"
                "• `ВОЗВРАТ 15.04` — возврат\n\n"
                "Или напишите сообщение со словом «возвратов» перед файлом.",
                parse_mode="Markdown"
            )

    except Exception as e:
        log.exception("Ошибка: %s", e)
        await msg.edit_text(f"❌ Ошибка при обработке файла:\n{str(e)[:200]}")
    finally:
        os.unlink(path)


async def on_topicid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    thread_id = update.message.message_thread_id or 0
    chat_id   = update.message.chat_id
    await update.message.reply_text(
        f"📌 *Информация о топике*\n\n"
        f"Chat ID: `{chat_id}`\n"
        f"Topic ID: `{thread_id}`\n\n"
        f"Добавьте в Render → Environment variables:\n\n"
        f"Если это топик *Приёмка* (отгрузки):\n"
        f"`TOPIC_PRIEMKA` = `{thread_id}`\n\n"
        f"Если это топик *Отгрузка* (возвраты):\n"
        f"`TOPIC_OTGRUZKA` = `{thread_id}`",
        parse_mode="Markdown"
    )


async def on_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        sh = db.table("shipments").select("project,status,parcels_count").execute().data
        rt = db.table("returns").select("status,orders_count,total_cost").execute().data

        def ss(proj):
            ps = [s for s in sh if s["project"] == proj]
            return (
                len(ps),
                sum(s["parcels_count"] for s in ps),
                sum(s["parcels_count"] for s in ps if s["status"] == "confirmed"),
                sum(s["parcels_count"] for s in ps if s["status"] == "new"),
            )

        ab, at, ac, ap = ss("ali")
        mb, mt, mc, mp = ss("mko")
        rb = len(rt)
        rc = sum(r["orders_count"] for r in rt if r["status"] == "confirmed")
        rs = sum(r["total_cost"] for r in rt)

        await update.message.reply_text(
            f"📊 *ShipTrack — Статистика*\n\n"
            f"🛒 *AliExpress* — {ab} партий\n"
            f"  Посылок: {at} | ✅ {ac} | ⏳ {ap}\n\n"
            f"📦 *Uzum Crossborder* — {mb} партий\n"
            f"  Посылок: {mt} | ✅ {mc} | ⏳ {mp}\n\n"
            f"🔄 *Возвраты* — {rb} актов\n"
            f"  ✅ Принято: {rc} | Сумма: {rs:,.0f} ₽",
            parse_mode="Markdown"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def on_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 *ShipTrack — Инструкция*\n\n"
        "В топике *Приёмка* отправьте файл с подписью:\n"
        "🛒 `АЛИ 15.04` — отгрузка AliExpress\n"
        "📦 `МКО` — отгрузка Uzum Crossborder\n\n"
        "В топике *Отгрузка* отправьте файл возврата — "
        "бот определит его автоматически.\n"
        "Форматы: .xlsx, .xls, .csv, .pdf\n\n"
        "/status — статистика\n"
        "/topicid — ID текущего топика\n"
        "/help — эта справка",
        parse_mode="Markdown"
    )


# ── Запуск с корректным перезапуском при ошибках ─────────────────────────────

async def run_bot():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    app.add_handler(MessageHandler(filters.Document.ALL, on_document))
    app.add_handler(CommandHandler("status",  on_status))
    app.add_handler(CommandHandler("topicid", on_topicid))
    app.add_handler(CommandHandler("help",    on_help))
    app.add_handler(CommandHandler("start",   on_help))
    async with app:
        await app.start()
        await app.updater.start_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True
        )
        log.info("🤖 ShipTrack Bot запущен")
        while True:
            await asyncio.sleep(1)


def main():
    while True:
        try:
            asyncio.run(run_bot())
        except Exception as e:
            log.error("❌ Бот упал с ошибкой: %s", e)
            log.info("🔄 Перезапуск через 5 секунд...")
            time.sleep(5)


if __name__ == "__main__":
    main()
